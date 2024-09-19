import os
import openpyxl
from openpyxl.xml.functions import QName, fromstring
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware
from typing import List

from app.routes.auth import auth_router
from app.routes.bundle import bundle_router
from app.routes.flashcard import flashcard_router
from app.core.error_handler import Error404
app = FastAPI()

origins: list[str] = [
    "http://localhost:3000"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

app.include_router(auth_router)
app.include_router(bundle_router)
app.include_router(flashcard_router)


@app.get("/download_file/{filename}")
async def download_file(filename: str):
    file_path = f"app/data/{filename}"
    if os.path.exists(file_path):
        return FileResponse(path=file_path, filename=filename)
    else:
        return {"error": "File not found"}


@app.get("/excel")
async def download_file():
    file_path = f"app/data/excel.xlsx"
    if not os.path.exists(file_path):
        raise Error404("File not found")

    file_size = os.path.getsize(file_path)

    def iterfile():
        with open(file_path, mode="rb") as file_like:
            yield from file_like

    headers = {
        "Content-Length": str(file_size),
        "Content-Disposition": f"attachment; filename=excel.xlsx"
    }

    return StreamingResponse(
        iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@app.get("/sheet_list", response_model=List[str])
async def get_sheet_names():
    file_path = f"app/data/excel.xlsx"
    if not os.path.exists(file_path):
        raise Error404("File not found")

    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    return sheet_names


def convert_to_hex_color(color_code):
    if len(color_code) == 8:
        rgb_hex = color_code[2:]
        hex_color = f"#{rgb_hex}"
        return hex_color
    else:
        return "#000000"


HLSMAX = 240
RGBMAX = 255


def get_theme_colors(wb):
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    theme_el = root.find(str(QName(xlmns, 'themeElements')))
    color_schemes = theme_el.findall(str(QName(xlmns, 'clrScheme')))
    first_color_scheme = color_schemes[0]

    colors = []
    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = first_color_scheme.find(str(QName(xlmns, c)))

        if accent.find(str(QName(xlmns, 'srgbClr'))) is not None:
            colors.append(accent.find(str(QName(xlmns, 'srgbClr'))).get('val'))
        elif accent.find(str(QName(xlmns, 'sysClr'))) is not None:
            colors.append(accent.find(str(QName(xlmns, 'sysClr'))).get('lastClr'))
        else:
            colors.append('000000')  # Default to black if color not found
    return colors


def tint_color(rgb, tint):
    # Convert hex to RGB
    r = int(rgb[0:2], 16)
    g = int(rgb[2:4], 16)
    b = int(rgb[4:6], 16)

    if tint < 0:
        r = int(r * (1.0 + tint))
        g = int(g * (1.0 + tint))
        b = int(b * (1.0 + tint))
    else:
        r = int(r * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint)))
        g = int(g * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint)))
        b = int(b * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint)))

    # Ensure values are in valid range
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))

    return f"{r:02X}{g:02X}{b:02X}"


def theme_and_tint_to_hex(theme_index, tint, wb):
    theme_colors = get_theme_colors(wb)
    rgb = theme_colors[theme_index]

    # Remove the alpha channel if present
    if len(rgb) == 8:  # e.g., 'FF0000FF'
        rgb = rgb[2:]  # 'FF0000'

    # Tint the RGB color
    hex_color = tint_color(rgb, tint)
    return f"#{hex_color}"


def get_fill_color(fill_color, wb):
    if fill_color.type == "rgb":
        return None if fill_color.rgb is None or fill_color.rgb == "00000000" else f"#{str(fill_color.rgb)[2:]}"
    elif fill_color.type == "theme":
        return theme_and_tint_to_hex(fill_color.theme, fill_color.tint, wb)
    else:
        return None


@app.get("/sheet_data/{sheet_name}")
async def get_sheet_data(sheet_name: str):
    file_path = "app/data/excel.xlsx"
    if not os.path.exists(file_path):
        raise Error404("File not found")

    try:
        # Load the workbook and the specific sheet
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        # Prepare data
        cell_data = []
        border_info = []

        print(wb.loaded_theme)

        for row_index, row in enumerate(sheet.iter_rows()):
            for col_index, cell in enumerate(row):
                # Extract cell value
                value = "<Image>" if cell.data_type == "e" else str(cell.value) if cell.value is not None else ""

                # Add cell data
                cell_data.append({
                    "r": row_index,
                    "c": col_index,
                    "v": {
                        "v": value,
                        "ff": cell.font.name,
                        "bl": 0 if cell.font.b is False else 1,
                        "it": 0 if cell.font.i is False else 1,
                        "fs": cell.font.size,
                        "ht": 0 if cell.alignment.horizontal == 'center' else 1 if cell.alignment.horizontal == 'left' else 2,
                        "fc": convert_to_hex_color(str(cell.font.color.rgb)),
                        "bg": get_fill_color(cell.fill.start_color, wb),
                    }
                })

                border_info.append({
                    "rangeType": "cell",
                    "value": {
                        "row_index": row_index,
                        "col_index": col_index,
                        "l":  None if cell.border.left.style is None else
                            {
                                "style": 7 if cell.border.left.style == "thin" else 8,
                                "color": "rgb(0, 0, 0)"
                            },
                        "r":  None if cell.border.left.style is None else
                            {
                                "style": 7 if cell.border.right.style == "thin" else 8,
                                "color": "rgb(0, 0, 0)"
                            },
                        "t":  None if cell.border.left.style is None else
                            {
                                "style": 7 if cell.border.top.style == "thin" else 8,
                                "color": "rgb(0, 0, 0)"
                            },
                        "b":  None if cell.border.left.style is None else
                            {
                                "style": 7 if cell.border.bottom.style == "thin" else 8,
                                "color": "rgb(0, 0, 0)"
                            },
                    }
                })

        sheet_data = {
            "name": sheet_name,
            "celldata": cell_data,
            "config": {
                "borderInfo": border_info
            },
        }

        return JSONResponse(content=sheet_data)
    except KeyError:
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
